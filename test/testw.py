import lenexdb
from lenexdb.baseapi import BaseApi, Athlete, Club, Event
import json
import openpyxl
import re
from datetime import time, datetime

import lenexdb.baseapi

names = json.load(open("data.json", "rb+"))
registered = dict(zip(names.values(), names.keys()))
lenexdb.baseapi.ENCODING = 'utf-8'


class RegisteredDistance:
    clubs = dict()
    athletes = dict()

    def __init__(self, xpath: str, table: str, config: dict):
        self.bapi = BaseApi(xpath, 'result')
        self.config = config
        workbook = openpyxl.load_workbook(table)
        sheet = workbook.active
        for i in sheet.iter_rows():
            row = tuple(r.value for r in i)
            break
        for row_k in sheet.iter_rows(min_row=2):
            row = tuple(r.value for r in row_k)
            if row[config['lastname']] is None:
                break
            club = self.get_club(row[self.config['club']])
            event = self.find_swimstyle(
                self.parse_gender(row[self.config['gender']]),
                registered.get(row[self.config['stroke']]),
                int(row[self.config['distance']])
            )
            athlete = self.get_athlete(club, event, row)
            et = self.parse_entrytime(row[self.config['entrytime']])
            athlete.add_entry(event.eventid, et)

    def get_lisense(self, event: Event, license: str | None) -> str | None:
        if license is None:
            return None
        for ts in event.timestandardrefs:
            if ts.marker == license:
                return ts.marker
        return None

    def parse_entrytime(self, entrytime) -> str:
        if isinstance(entrytime, (time, datetime)):
            r = entrytime.strftime('00:%H:%M.%S')
            return r
        m = re.fullmatch("(\d{2,3}):(\d{2}):(\d{2})", entrytime)
        if m is None:
            return "00:00:00.00"
        return f"00:{m.group(1)}:{m.group(2)}.{m.group(3)}"

    def find_swimstyle(self, gdr, srk, dist) -> Event:
        for e in self.bapi.events:
            if (
                e.gender == gdr
                and e.swim_style.stroke == srk
                and e.swim_style.distance == dist
            ):
                return e
        raise TypeError("Incorrect distance")

    def parse_gender(self, gender: str):
        if gender == "Мужской":
            return "M"
        if gender == "Женский":
            return "F"
        raise

    def parse_bd(self, bd: str):
        if not isinstance(bd, str):
            return bd
        dbl = bd.split(".")
        dbl.reverse()
        return "-".join(dbl)

    def get_club(self, name: str) -> Club:
        if name.lower() not in self.clubs:
            self.clubs[name.lower()] = self.bapi.create_club(name)
        return self.clubs[name.lower()]

    def get_athlete(self, club: Club, event: Event, row: tuple) -> Athlete:
        key = row[self.config['lastname']] + \
            " " + row[self.config['firstname']]
        if key not in self.athletes:
            athl = club.create_athlete(
                row[self.config['lastname']],
                row[self.config['firstname']],
                self.parse_bd(row[self.config['birthdate']]),
                self.parse_gender(row[self.config['gender']]),
                self.get_lisense(event, row[self.config['license']]),
            )
            self.athletes[key] = athl
        return self.athletes[key]


# config = {
#     'lastname': 0,
#     'firstname': 1,
#     'birthdate': 5,
#     'gender': 3,
#     'license':  4,
#     'club': 8,
#     'stroke': 10,
#     'distance': 9,
#     'entrytime': 12
# }

config = {
    'lastname': 0,
    'firstname': 1,
    'gender': 2,
    'license':  3,
    'birthdate': 4,
    'club': 5,
    'stroke': 6,
    'distance': 7,
    'entrytime': 8
}
xpath = r"C:\Users\2008d\Downloads\Telegram Desktop\20250209_Lenex.lxf"
rd = RegisteredDistance(xpath, "test.xlsx", config)
rd.bapi.save("result/test.lef")
print(rd)
