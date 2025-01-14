from lenexdb.baseapi import BaseApi, Athlete, Club, Event, AgeGroup
from lenexdb.basetime import BaseTime
import openpyxl
import re
import time
from datetime import time as dtime, datetime, date
from typing import List
import logging
import math
from functools import cache
try:
    bt = BaseTime()
except FileNotFoundError:
    bt = BaseTime.null()


def get_age(s: str):
    today = date.today()
    born = datetime.strptime(s, "%Y-%m-%d")
    return today.year - born.year
    # return today.year - born.year - ((today.month, today.day) < (born.month, born.day))


@cache
def check_age(age: int, minage: int, maxage: int):
    if maxage == -1:
        return True
    return minage <= age <= maxage


def sum_age_groups(agegroups: List[AgeGroup]):
    amax, amin = -1, math.inf
    for group in agegroups:
        if group.agemax > amax or group.agemax == -1:
            amax = group.agemax
        if group.agemin < amin:
            amin = group.agemin
    return amin, amax


def get_only_time(entrytime: str):
    hrs, mins, secs = entrytime.split(':')
    return int(hrs)*60*60+int(mins)*60+float(secs)


class RegisteredDistance:
    clubs: dict[str, Club] = dict()
    athletes: dict[str, Athlete] = dict()
    logger = logging.getLogger()
    debug = True
    delay = 0.01
    basetime = bt

    def __init__(self, xpath: str, table: str, data: dict, **kwargs):
        self.__dict__.update(kwargs)

        self.bapi = BaseApi(xpath)
        self.data = data
        self.config = data['location']
        self.registered = data['reversed_styles']

        workbook = openpyxl.load_workbook(table)
        self.sheet = workbook.active

    @classmethod
    def init(cls) -> 'RegisteredDistance':
        return cls.__new__(cls)

    @classmethod
    def base_init(cls, baseapi: BaseApi, table: str, data: dict | None = None) -> 'RegisteredDistance':
        self = cls.init()

        self.bapi = baseapi

        workbook = openpyxl.load_workbook(table)
        self.sheet = workbook.active

        if data is not None:
            self.data = data
            self.config = data['location']
            self.registered = data['reversed_styles']

        return self

    def add_elements(self):
        for club in self.bapi.clubs:
            self.clubs[club.name.lower()] = club
        for athlete in self.bapi.athletes:
            key = athlete.lastname + \
                " " + athlete.firstname
            self.athletes[key.lower()] = athlete

    def parse(self):
        for row_k in self.sheet.iter_rows(min_row=2):
            row = tuple(r.value for r in row_k)
            if row[self.config['lastname']] is None:
                break
            if self.debug:
                self.logger.debug(
                    f'Parse {row_k[0].row} row: '+' '.join(map(str, row)))
            if self.delay:
                time.sleep(self.delay)
            club = self.get_club(row[self.config['club']])
            athlete = self.get_athlete(club, row)
            try:
                event = self.find_swimstyle(
                    athlete.gender,
                    self.registered.get(row[self.config['stroke']]),
                    int(row[self.config['distance']]),
                    self.parse_bd(row[self.config['birthdate']], age=True)
                )
            except TypeError:
                self.logger.error(
                    f"Missed the distance {row[self.config['distance']]} {row[self.config['stroke']]}")
                continue
            et = self.parse_entrytime(row[self.config['entrytime']])
            if self.data['points']['switch']:
                point = self.basetime.get_point(
                    athlete.gender,
                    int(row[self.config['distance']]),
                    self.registered.get(row[self.config['stroke']]),
                    get_only_time(et)
                )
                et, result = self.validate_entry_time(et, point)
                if result:
                    self.logger.warning(
                        f"{athlete.firstname} {athlete.lastname} Wrong time! Points: {point:.5f}, Entry Time: {row[self.config['entrytime']]}")
            athlete.add_entry(event.eventid, et)

    def validate_entry_time(self, entrytime: str, point: float) -> tuple[str, bool]:
        points = self.data.get('points')
        min, max = points['min'], points['max']
        if point == 0:
            return entrytime, False
        if max > point > min:
            return entrytime, False
        return '00:00:00:00', True

    @cache
    def parse_lisense(self, license: str) -> str:
        def chg(match: re.Match):
            t = match.string[match.regs[0][0]:match.regs[0][1]]
            return t.lower()
        license = re.sub('[^I]{1}', chg, license)

        for a, b in self.data['replacement'].items():
            license = license.replace(a, b)
        return license

    @cache
    def get_lisense(self, license: str | None) -> str | None:
        if license is None:
            return None
        license = self.parse_lisense(str(license))
        if license in self.data['lisenses']:
            return license
        self.logger.error('Not found license ' + license)

    @cache
    def parse_entrytime(self, entrytime) -> str:
        if entrytime is None:
            return "00:00:00.00"
        if isinstance(entrytime, (dtime, datetime)):
            r = entrytime.strftime('00:%H:%M.%S')
            return r
        m = re.fullmatch("(\d{1,3}):(\d{2}):(\d{1,2})", entrytime)
        if m is None:
            self.logger.error('Incorrect entrytime ' + entrytime)
            return "00:00:00.00"
        return f"00:{m.group(1)}:{m.group(2)}.{m.group(3)}"

    @cache
    def find_swimstyle(self, gdr, srk, dist, age) -> Event:
        for e in self.bapi.events:
            amin, amax = sum_age_groups(e.agegroups)
            if (
                e.gender == gdr
                and e.swim_style.stroke == srk
                and e.swim_style.distance == dist
                and check_age(age, amin, amax)
            ):
                return e
        self.logger.error(f'Incorrect distance {gdr} {dist} {srk} {age}')
        raise TypeError("Incorrect distance")

    @cache
    def parse_gender(self, gender: str):
        if gender == "Мужской":
            return "M"
        if gender == "Женский":
            return "F"
        self.logger.error(f'Incorrect gender {gender}')
        raise TypeError("Sex not found")

    @cache
    def parse_bd(self, bd: str, age: bool):
        if not isinstance(bd, str):
            return bd
        dbl = bd.split(".")
        dbl.reverse()
        d = "-".join(dbl)
        return get_age(d) if age else d

    def get_club(self, name: str) -> Club:
        if name.lower() not in self.clubs:
            self.clubs[name.lower()] = self.bapi.create_club(name)
        return self.clubs[name.lower()]

    def get_athlete(self, club: Club, row: tuple) -> Athlete:
        key = (row[self.config['lastname']]
               + " " + row[self.config['firstname']]).lower()
        if key not in self.athletes:
            athl = club.create_athlete(
                row[self.config['lastname']],
                row[self.config['firstname']],
                self.parse_bd(row[self.config['birthdate']], age=False),
                self.parse_gender(row[self.config['gender']]),
                self.get_lisense(row[self.config['license']]),
            )
            self.athletes[key] = athl
        return self.athletes[key]


if __name__ == '__main__':
    # data = json.load(open("data.json", "rb+"))
    # xpath = r"C:\Users\2008d\Downloads\Telegram Desktop\20250209_Lenex.lxf"
    # rd = RegisteredDistance(xpath, "test3.xlsx", data)
    # rd.bapi.save("result/test.lxf")
    # print(rd)

    # workbook = openpyxl.load_workbook('test.xlsx')
    # sheet = workbook.active
    # values = [r.value for r in sheet[1]]
    # d = dict(zip(values, range(len(values))))
    # print(json.dumps(d, ensure_ascii=False))

    agegroups = [
        AgeGroup(None, 0, 10, 9),
        AgeGroup(None, 0, 8, 8),
    ]
    agegroups2 = [
        AgeGroup(None, 0, 13, 11),
        AgeGroup(None, 0, 15, 14),
        AgeGroup(None, 0, -1, 16),
    ]
    print(sum_age_groups(agegroups))
    print(sum_age_groups(agegroups2))
